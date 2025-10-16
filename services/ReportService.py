import pandas as pd
from io import BytesIO
from sqlalchemy import select

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from models import (
    User,
    Object,
    Speciality,
    Attendance
)

class ReportService:
    def __init__(self, session):
        self.session = session

    async def generate_report(self):
        data = await self._fetch_data()
        buffer = self._save_to_excel(data)
        self._format_excel(buffer)
        return buffer

    async def _fetch_data(self):
        stmt = (
            select(
                Attendance.date.label("Дата"),
                User.telegram_id.label("Telegram ID"),
                (User.last_name + " " + User.first_name + " " +
                 (User.middle_name if User.middle_name else "")).label("ФИО"),
                User.phone.label("Телефон"),
                Speciality.name.label("Специальность"),
                Attendance.action.label("Действие"),
                Object.name.label("Объект"),
                Attendance.note.label("Примичание")
            )
            .select_from(Attendance)
            .join(User, Attendance.user_id == User.id)
            .join(Object, Attendance.object_id == Object.id)
            .join(Speciality, User.speciality_id == Speciality.id)
            .order_by(Attendance.date)
        )

        result = await self.session.execute(stmt)

        return result.all()

    @staticmethod
    def _save_to_excel(data: list[tuple]):
        df = pd.DataFrame(data, columns = [
            "Дата",
            "Telegram ID",
            "ФИО",
            "Телефон",
            "Специальность",
            "Действие",
            "Объект",
            "Примичание"
        ])
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine = "openpyxl") as writer: # type: ignore
            df.to_excel(writer, index = False, sheet_name = "Attendance")
        buffer.seek(0)

        return buffer

    @staticmethod
    def _format_excel(buffer: BytesIO):
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active

        header_font = Font(bold = True)
        center_align = Alignment(horizontal = "center", vertical = "center")
        thin_border = Border(
            left = Side(style = 'thin'),
            right = Side(style = 'thin'),
            top = Side(style = 'thin'),
            bottom = Side(style = 'thin')
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for row in ws.iter_rows(min_row = 2):
            for cell in row:
                cell.alignment = Alignment(vertical = "center")
                cell.border = thin_border

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column].width = max_length + 2

        new_buffer = BytesIO()
        wb.save(new_buffer)
        new_buffer.seek(0)
        buffer.seek(0)

        return new_buffer